"""
產生預設 Excel 模板檔案

每個模板包含:
- 標題列
- 基本資訊區（設備名稱、編號、日期、檢查人員）
- 檢查項目（值/判定/備註欄）
- 簽核區
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def _apply_styles(ws, title_row=1):
    """共用樣式設定"""
    # 欄寬
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 20

    # 標題樣式
    title_font = Font(name="Microsoft JhengHei", size=14, bold=True)
    header_font = Font(name="Microsoft JhengHei", size=10, bold=True)
    normal_font = Font(name="Microsoft JhengHei", size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    return title_font, header_font, normal_font, header_fill, thin_border, center_align, left_align


def generate_electrical_inspection():
    """電氣設備定期檢查表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "電氣設備定檢"
    title_font, header_font, normal_font, header_fill, thin_border, center_align, left_align = _apply_styles(ws)

    # 標題
    ws.merge_cells('A1:F1')
    ws['A1'] = "電氣設備定期檢查表（通用版）"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # 基本資訊
    info_rows = [
        ("設備名稱：", "", "設備編號：", ""),
        ("檢查日期：", "", "檢查人員：", ""),
        ("設備位置：", "", "設備類型：", ""),
    ]
    for i, (l1, v1, l2, v2) in enumerate(info_rows, start=3):
        ws.merge_cells(f'A{i}:B{i}')
        ws[f'A{i}'] = l1
        ws[f'A{i}'].font = header_font
        ws[f'C{i}'] = v1
        ws.merge_cells(f'D{i}:E{i}')
        ws[f'D{i}'] = l2
        ws[f'D{i}'].font = header_font
        ws[f'F{i}'] = v2

    # 檢查項目標題
    row = 7
    headers = ["序號", "檢查項目", "量測值", "判定", "標準值", "備註"]
    cols = ['A', 'B', 'C', 'D', 'E', 'F']
    for c, h in zip(cols, headers):
        ws[f'{c}{row}'] = h
        ws[f'{c}{row}'].font = header_font
        ws[f'{c}{row}'].fill = header_fill
        ws[f'{c}{row}'].border = thin_border
        ws[f'{c}{row}'].alignment = center_align

    # 檢查項目
    items = [
        ("1", "絕緣電阻", "", "", "≥1.0 MΩ", ""),
        ("2", "接地電阻", "", "", "≤100 Ω", ""),
        ("3", "漏電斷路器動作時間", "", "", "≤100 ms", ""),
        ("4", "電壓", "", "", "220±10% V", ""),
        ("5", "電流", "", "", "依額定", ""),
        ("6", "接線端子鬆緊度", "", "", "正常/異常", ""),
        ("7", "外觀及清潔狀況", "", "", "良好/不良", ""),
        ("8", "指示燈功能", "", "", "正常/異常", ""),
    ]
    for j, item in enumerate(items):
        r = row + 1 + j
        for c, val in zip(cols, item):
            ws[f'{c}{r}'] = val
            ws[f'{c}{r}'].font = normal_font
            ws[f'{c}{r}'].border = thin_border
            ws[f'{c}{r}'].alignment = center_align if c in ('A', 'D', 'E') else left_align

    # 簽核區
    sig_row = row + 1 + len(items) + 2
    ws.merge_cells(f'A{sig_row}:F{sig_row}')
    ws[f'A{sig_row}'] = "簽核區"
    ws[f'A{sig_row}'].font = header_font
    ws[f'A{sig_row}'].alignment = center_align

    ws[f'A{sig_row+1}'] = "檢查人員簽名："
    ws[f'A{sig_row+1}'].font = normal_font
    ws[f'D{sig_row+1}'] = "主管簽核："
    ws[f'D{sig_row+1}'].font = normal_font

    return wb


def generate_fire_safety_inspection():
    """消防安全設備檢查表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "消防安全檢查"
    title_font, header_font, normal_font, header_fill, thin_border, center_align, left_align = _apply_styles(ws)

    ws.merge_cells('A1:F1')
    ws['A1'] = "消防安全設備檢查表（通用版）"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    info_rows = [
        ("設備名稱：", "", "設備編號：", ""),
        ("檢查日期：", "", "檢查人員：", ""),
        ("設備位置：", "", "設備類型：", ""),
    ]
    for i, (l1, v1, l2, v2) in enumerate(info_rows, start=3):
        ws.merge_cells(f'A{i}:B{i}')
        ws[f'A{i}'] = l1
        ws[f'A{i}'].font = header_font
        ws[f'C{i}'] = v1
        ws.merge_cells(f'D{i}:E{i}')
        ws[f'D{i}'] = l2
        ws[f'D{i}'].font = header_font
        ws[f'F{i}'] = v2

    row = 7
    headers = ["序號", "檢查項目", "量測值", "判定", "標準值", "備註"]
    cols = ['A', 'B', 'C', 'D', 'E', 'F']
    for c, h in zip(cols, headers):
        ws[f'{c}{row}'] = h
        ws[f'{c}{row}'].font = header_font
        ws[f'{c}{row}'].fill = header_fill
        ws[f'{c}{row}'].border = thin_border
        ws[f'{c}{row}'].alignment = center_align

    items = [
        ("1", "滅火器壓力", "", "", "綠色區域", ""),
        ("2", "滅火器有效期限", "", "", "未過期", ""),
        ("3", "消防栓水壓", "", "", "≥2.0 kgf/cm²", ""),
        ("4", "消防栓外觀", "", "", "良好/不良", ""),
        ("5", "火警受信總機", "", "", "正常/異常", ""),
        ("6", "偵煙探測器", "", "", "正常/異常", ""),
        ("7", "緊急照明燈", "", "", "正常/異常", ""),
        ("8", "避難方向指示燈", "", "", "正常/異常", ""),
    ]
    for j, item in enumerate(items):
        r = row + 1 + j
        for c, val in zip(cols, item):
            ws[f'{c}{r}'] = val
            ws[f'{c}{r}'].font = normal_font
            ws[f'{c}{r}'].border = thin_border
            ws[f'{c}{r}'].alignment = center_align if c in ('A', 'D', 'E') else left_align

    sig_row = row + 1 + len(items) + 2
    ws.merge_cells(f'A{sig_row}:F{sig_row}')
    ws[f'A{sig_row}'] = "簽核區"
    ws[f'A{sig_row}'].font = header_font
    ws[f'A{sig_row}'].alignment = center_align
    ws[f'A{sig_row+1}'] = "檢查人員簽名："
    ws[f'A{sig_row+1}'].font = normal_font
    ws[f'D{sig_row+1}'] = "主管簽核："
    ws[f'D{sig_row+1}'].font = normal_font

    return wb


def generate_motor_inspection():
    """馬達/泵浦定期檢查表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "馬達泵浦定檢"
    title_font, header_font, normal_font, header_fill, thin_border, center_align, left_align = _apply_styles(ws)

    ws.merge_cells('A1:F1')
    ws['A1'] = "馬達/泵浦定期檢查表"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    info_rows = [
        ("設備名稱：", "", "設備編號：", ""),
        ("檢查日期：", "", "檢查人員：", ""),
        ("設備位置：", "", "設備類型：", ""),
    ]
    for i, (l1, v1, l2, v2) in enumerate(info_rows, start=3):
        ws.merge_cells(f'A{i}:B{i}')
        ws[f'A{i}'] = l1
        ws[f'A{i}'].font = header_font
        ws[f'C{i}'] = v1
        ws.merge_cells(f'D{i}:E{i}')
        ws[f'D{i}'] = l2
        ws[f'D{i}'].font = header_font
        ws[f'F{i}'] = v2

    row = 7
    headers = ["序號", "檢查項目", "量測值", "判定", "標準值", "備註"]
    cols = ['A', 'B', 'C', 'D', 'E', 'F']
    for c, h in zip(cols, headers):
        ws[f'{c}{row}'] = h
        ws[f'{c}{row}'].font = header_font
        ws[f'{c}{row}'].fill = header_fill
        ws[f'{c}{row}'].border = thin_border
        ws[f'{c}{row}'].alignment = center_align

    items = [
        ("1", "振動值（軸承端）", "", "", "≤4.5 mm/s", ""),
        ("2", "軸承溫度", "", "", "≤80 ℃", ""),
        ("3", "運轉電流", "", "", "依額定", ""),
        ("4", "絕緣電阻", "", "", "≥1.0 MΩ", ""),
        ("5", "噪音", "", "", "≤85 dB", ""),
        ("6", "潤滑油位/油質", "", "", "正常/異常", ""),
        ("7", "聯軸器對心", "", "", "正常/異常", ""),
        ("8", "外觀及清潔狀況", "", "", "良好/不良", ""),
    ]
    for j, item in enumerate(items):
        r = row + 1 + j
        for c, val in zip(cols, item):
            ws[f'{c}{r}'] = val
            ws[f'{c}{r}'].font = normal_font
            ws[f'{c}{r}'].border = thin_border
            ws[f'{c}{r}'].alignment = center_align if c in ('A', 'D', 'E') else left_align

    sig_row = row + 1 + len(items) + 2
    ws.merge_cells(f'A{sig_row}:F{sig_row}')
    ws[f'A{sig_row}'] = "簽核區"
    ws[f'A{sig_row}'].font = header_font
    ws[f'A{sig_row}'].alignment = center_align
    ws[f'A{sig_row+1}'] = "檢查人員簽名："
    ws[f'A{sig_row+1}'].font = normal_font
    ws[f'D{sig_row+1}'] = "主管簽核："
    ws[f'D{sig_row+1}'].font = normal_font

    return wb


def generate_5s_audit():
    """廠區 5S 巡查表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "5S巡查"
    title_font, header_font, normal_font, header_fill, thin_border, center_align, left_align = _apply_styles(ws)

    ws.merge_cells('A1:F1')
    ws['A1'] = "廠區 5S 巡查表"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    info_rows = [
        ("巡查區域：", "", "區域編號：", ""),
        ("巡查日期：", "", "巡查人員：", ""),
    ]
    for i, (l1, v1, l2, v2) in enumerate(info_rows, start=3):
        ws.merge_cells(f'A{i}:B{i}')
        ws[f'A{i}'] = l1
        ws[f'A{i}'].font = header_font
        ws[f'C{i}'] = v1
        ws.merge_cells(f'D{i}:E{i}')
        ws[f'D{i}'] = l2
        ws[f'D{i}'].font = header_font
        ws[f'F{i}'] = v2

    row = 6
    headers = ["序號", "檢查項目", "評分(1-5)", "判定", "標準", "備註"]
    cols = ['A', 'B', 'C', 'D', 'E', 'F']
    for c, h in zip(cols, headers):
        ws[f'{c}{row}'] = h
        ws[f'{c}{row}'].font = header_font
        ws[f'{c}{row}'].fill = header_fill
        ws[f'{c}{row}'].border = thin_border
        ws[f'{c}{row}'].alignment = center_align

    items = [
        ("1", "整理：不需要物品是否已移除", "", "", "≥3分", ""),
        ("2", "整頓：工具/物料定位標示", "", "", "≥3分", ""),
        ("3", "清掃：地面/設備清潔度", "", "", "≥3分", ""),
        ("4", "清潔：標準化維持狀況", "", "", "≥3分", ""),
        ("5", "素養：人員遵守規定狀況", "", "", "≥3分", ""),
        ("6", "通道暢通", "", "", "合格/不合格", ""),
        ("7", "安全標示完整性", "", "", "合格/不合格", ""),
        ("8", "廢棄物分類", "", "", "合格/不合格", ""),
    ]
    for j, item in enumerate(items):
        r = row + 1 + j
        for c, val in zip(cols, item):
            ws[f'{c}{r}'] = val
            ws[f'{c}{r}'].font = normal_font
            ws[f'{c}{r}'].border = thin_border
            ws[f'{c}{r}'].alignment = center_align if c in ('A', 'D', 'E') else left_align

    sig_row = row + 1 + len(items) + 2
    ws.merge_cells(f'A{sig_row}:F{sig_row}')
    ws[f'A{sig_row}'] = "簽核區"
    ws[f'A{sig_row}'].font = header_font
    ws[f'A{sig_row}'].alignment = center_align
    ws[f'A{sig_row+1}'] = "巡查人員簽名："
    ws[f'A{sig_row+1}'].font = normal_font
    ws[f'D{sig_row+1}'] = "主管簽核："
    ws[f'D{sig_row+1}'].font = normal_font

    return wb


def generate_all():
    """產生所有預設模板"""
    out_dir = os.path.dirname(__file__)

    templates = {
        "electrical_inspection.xlsx": generate_electrical_inspection,
        "fire_safety_inspection.xlsx": generate_fire_safety_inspection,
        "motor_inspection.xlsx": generate_motor_inspection,
        "5s_audit.xlsx": generate_5s_audit,
    }

    for filename, gen_fn in templates.items():
        wb = gen_fn()
        path = os.path.join(out_dir, filename)
        wb.save(path)
        print(f"Generated: {path}")


if __name__ == "__main__":
    generate_all()
